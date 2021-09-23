import telebot
from telebot import types
import openpyxl
from settings import API_KEY

#Bot Initialization
bot = telebot.TeleBot(API_KEY)
bot.SESSION_TIME_TO_LIVE = 60 * 5

#Initializing Keyboard Markup buttons text
optionA = 'Search University by name or City'
optionB = 'Search Universities offering application fee waiver'
optionC = 'Search by name or city'
optionD = 'All Universities'
optionE = 'Go Back'

#Initializing Keyboard Markup buttons
markup = types.ReplyKeyboardMarkup(row_width=1, one_time_keyboard=True)
itembtn1 = types.KeyboardButton(optionA)
itembtn2 = types.KeyboardButton(optionB)
markup.add(itembtn1, itembtn2)

markup1 = types.ReplyKeyboardMarkup(row_width=1, one_time_keyboard=True)
itembtn3 = types.KeyboardButton(optionC)
itembtn4 = types.KeyboardButton(optionD)
itembtn5 = types.KeyboardButton(optionE)
markup1.add(itembtn3, itembtn4, itembtn5)

#Start of the program
@bot.message_handler(commands=['start'])
def start(message):
  reply = bot.send_message(message.chat.id, 'Hello, welcome!\nHere, you can find Application & Tution fee details of various Universities.\nChoose an option to continue: ', reply_markup=markup)
  bot.register_next_step_handler(reply, check)

@bot.message_handler(commands=['continue'])
def cont(message):
  reply = bot.send_message(message.chat.id,'Choose an option to continue: ', reply_markup=markup)
  bot.register_next_step_handler(reply, check)

#Checking the option selected in first page
def check(message):
  if message.text == optionA:
    UniversitySearch(message)
  elif message.text == optionB:
    afwUni(message)
  else:
    bot.send_message(message.chat.id, 'Wrong option selected, please try again')
    cont(message)
    
#Checking the option selected in second page    
def check2(message):
  if message.text == optionC:
    reply2 = bot.send_message(message.chat.id, 'Please enter the University name or city')
    bot.register_next_step_handler(reply2, afwUniversitySearch)
  elif message.text == optionD:
    bot.send_message(message.chat.id, 'Showing results for all the Universities with Application fee waiver')
    afwAll(message)
  elif message.text == optionE:
    cont(message)
  else:
    bot.send_message(message.chat.id, 'Wrong option selected, please try again')
    afwUni(message)
    
#Function to capture user input for university name or city
def UniversitySearch(message):
  msg = bot.send_message(message.chat.id, 'Please enter the University name or city')
  bot.register_next_step_handler(msg, universitySearch)

#Function which displays page two options
def afwUni(message):
  reply2 = bot.send_message(message.chat.id, 'Choose one below: ', reply_markup=markup1)
  bot.register_next_step_handler(reply2, check2)

#Function to capture user input and change it to lower text
def constants(message):
  collegeByUser = message.text
  collegeByUserLower = collegeByUser.lower()
  return collegeByUserLower

#Function to initialize excel file
def excel():
  wb = openpyxl.load_workbook('USUniDetails.xlsx')
  MainSheet = wb['Sheet1']
  return MainSheet

#University search by name or city
def universitySearch(message):
  MainSheet = excel()
  collegeByUserLower = constants(message)
  collegeName = list()
  count = 0
  for i in range(1, 329):
    collegeName.append(MainSheet.cell(row=i, column=2).value)
    converted_list = [x.lower() for x in collegeName]
    if any(collegeByUserLower in word for word in converted_list):
      CollegeNameFinal = ''.join(collegeName)
      ApplicationFee = str(MainSheet.cell(row=i, column=3).value)
      TutionFee = str(MainSheet.cell(row=i, column=4).value)

      bot.send_message(message.chat.id, 'College Name: ' + CollegeNameFinal + '\nApplication Fee: ' + ApplicationFee + '\nTution Fee/Year: ' + TutionFee)
      count = count + 1

    collegeName = list()
  endMessage(message, count)

#AFW university search by name or city
def afwUniversitySearch(message):
  MainSheet = excel()
  collegeByUserLower = constants(message)
  collegeName = list()
  count = 0
  for i in range(1, 329):
    collegeName.append(MainSheet.cell(row=i, column=2).value)
    converted_list = [x.lower() for x in collegeName]
    ApplicationFee = str(MainSheet.cell(row=i, column=3).value)
    TutionFee = str(MainSheet.cell(row=i, column=4).value)
    if ApplicationFee == 'AFW' or ApplicationFee == 'Free':
      if any(collegeByUserLower in word for word in converted_list):
        CollegeNameFinal = ''.join(collegeName)

        bot.send_message(message.chat.id, 'College Name: ' + CollegeNameFinal + '\nApplication Fee: ' + ApplicationFee + '\nTution Fee/Year: ' + TutionFee)
        count = count + 1
    collegeName = list()
  endMessage(message, count)

#AFW all universities
def afwAll(message):
  MainSheet = excel()
  collegeName = list()
  count = 0
  for i in range(1, 329):
    collegeName.append(MainSheet.cell(row=i, column=2).value)
    ApplicationFee = str(MainSheet.cell(row=i, column=3).value)
    TutionFee = str(MainSheet.cell(row=i, column=4).value)
    if ApplicationFee == 'AFW' or ApplicationFee == 'Free':
      CollegeNameFinal = ''.join(collegeName)

      bot.send_message(message.chat.id, 'College Name: ' + CollegeNameFinal + '\nApplication Fee: ' + ApplicationFee + '\nTution Fee/Year: ' + TutionFee)
      count = count + 1
    collegeName = list()
  endMessage(message, count)

#End message
def endMessage(message, count):
  if count == 0:
    bot.send_message(message.chat.id, 'No details found')
    bot.send_message(message.chat.id, 'For more information, contact a consultancy near you or check the University website.\nTo continue searching for universities, tap /continue')
  else:
    bot.send_message(message.chat.id, 'This is an estimate and is not guaranteed. For more details, contact a consultancy near you or check the University website.\nTo continue searching for universities, tap /continue')

  bot.register_next_step_handler(message, endError)

#Wrong input error
def endError(message):
  if message.text == '/continue':
    cont(message)
  elif message.text == '/start':
    start(message)
  else:
    bot.send_message(message.chat.id, 'Wrong input, please try again. \nTo continue searching for universities, tap /continue')
    bot.register_next_step_handler(message, endError)

bot.polling()
